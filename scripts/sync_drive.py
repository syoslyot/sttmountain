"""
從 Google Drive 同步出隊資料到本機，並產生 data/raw/sync_meta.json。

Drive 結構（2026-05-01 後的新格式）：
  所有出隊資料夾/
    {出隊名稱}/               ← solo：資料夾直接含直企 xlsx
      {日期}_{名稱}_直企.xlsx
      地圖/                   ← pdf, docx, jpg, png, jpeg
      航跡/                   ← gpx, kml
      上繳紀錄/               ← txt, md, docx, pdf, Google Doc
    {活動名稱}/               ← 大眾化：資料夾含子資料夾
      {隊伍名稱}/
        {日期}_{名稱}_直企.xlsx
        地圖/
        航跡/
        上繳紀錄/

環境變數：
  GDRIVE_CREDENTIALS_JSON  — Service Account JSON 內容
  GDRIVE_ROOT_FOLDER_ID    — 「所有出隊資料夾」的 Drive folder ID
  SUPABASE_URL             — Supabase project URL
  SUPABASE_SERVICE_KEY     — service_role key
"""

import io
import json
import os
import sys
from datetime import date, datetime, timezone
from pathlib import Path

from dotenv import load_dotenv
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from supabase import create_client

load_dotenv()
load_dotenv(os.environ.get("ENV_FILE", ".env.local"), override=True)

# ── 常數 ────────────────────────────────────────────────────

IMPORT_CUTOFF_DATE = date(2026, 5, 15)

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

RAW_DIR  = Path(__file__).parent.parent / "data" / "raw"
XLSX_DIR = RAW_DIR / "xlsx"
TXT_DIR  = RAW_DIR / "txt"
GPX_DIR  = Path(__file__).parent.parent / "app" / "static" / "gpx"
MAPS_DIR = Path(__file__).parent.parent / "app" / "static" / "maps"
META_PATH = RAW_DIR / "sync_meta.json"

MAP_FOLDER_NAMES    = {"地圖"}
TRACK_FOLDER_NAMES  = {"航跡"}
RECORD_FOLDER_NAMES = {"上繳紀錄"}

ZHIJIAN_EXTS = {".xlsx", ".xls", ".numbers"}
GPX_EXTS    = {".gpx", ".kml"}
MAP_EXTS    = {".pdf", ".docx", ".jpg", ".jpeg", ".png"}
RECORD_EXTS = {".txt", ".md", ".docx", ".pdf"}

FOLDER_MIME = "application/vnd.google-apps.folder"
GDOC_MIME   = "application/vnd.google-apps.document"
GSHEET_MIME = "application/vnd.google-apps.spreadsheet"


# ── Drive helpers ────────────────────────────────────────────

def build_service():
    cred_json = os.environ.get("GDRIVE_CREDENTIALS_JSON")
    if not cred_json:
        raise RuntimeError("GDRIVE_CREDENTIALS_JSON not set")
    creds = Credentials.from_service_account_info(json.loads(cred_json), scopes=SCOPES)
    return build("drive", "v3", credentials=creds)


def list_folder(service, folder_id: str) -> list[dict]:
    results, page_token = [], None
    while True:
        resp = service.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            fields="nextPageToken, files(id, name, mimeType, modifiedTime, createdTime)",
            pageToken=page_token,
        ).execute()
        results.extend(resp.get("files", []))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return results


def parse_dt(s: str) -> datetime:
    return datetime.fromisoformat(s.replace("Z", "+00:00"))


def is_folder(item: dict) -> bool:
    return item["mimeType"] == FOLDER_MIME


def _drive_newer(modified_time: str | None, dest: Path) -> bool:
    if not dest.exists():
        return True
    if not modified_time:
        return False
    return parse_dt(modified_time) > datetime.fromtimestamp(dest.stat().st_mtime, tz=timezone.utc)


def download_file(service, file_id: str, dest: Path, modified_time: str | None = None):
    dest.parent.mkdir(parents=True, exist_ok=True)
    if not _drive_newer(modified_time, dest):
        return
    request = service.files().get_media(fileId=file_id)
    with io.FileIO(dest, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
    print(f"  downloaded: {dest.name}")


def download_google_sheet_as_xlsx(service, file_id: str, dest: Path, modified_time: str | None = None) -> Path:
    dest = dest.with_suffix(".xlsx")
    dest.parent.mkdir(parents=True, exist_ok=True)
    if not _drive_newer(modified_time, dest):
        return dest
    request = service.files().export_media(
        fileId=file_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    with io.FileIO(dest, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
    print(f"  downloaded (gsheet→xlsx): {dest.name}")
    return dest


def convert_numbers_to_xlsx(src: Path) -> Path:
    import numbers_parser
    import openpyxl as xl

    dest = src.with_suffix(".xlsx")
    doc = numbers_parser.Document(str(src))
    wb = xl.Workbook()
    wb.remove(wb.active)

    for sheet in doc.sheets:
        ws = wb.create_sheet(title=sheet.name)
        row_offset = 0
        for table in sheet.tables:
            for cell in table.iter_rows():
                for c in cell:
                    if c.value is not None:
                        ws.cell(row=row_offset + c.row + 1, column=c.col + 1, value=c.value)
            row_offset += table.num_rows + 1

    wb.save(dest)
    print(f"  converted (numbers→xlsx): {dest.name}")
    return dest


def download_google_doc(service, file_id: str, dest: Path, modified_time: str | None = None):
    dest.parent.mkdir(parents=True, exist_ok=True)
    if not _drive_newer(modified_time, dest):
        return
    request = service.files().export_media(fileId=file_id, mimeType="text/plain")
    with io.FileIO(dest, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
    print(f"  downloaded (gdoc→txt): {dest.name}")


# ── Supabase helpers ─────────────────────────────────────────

def _supabase_client():
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_KEY")
    if not url or not key:
        return None
    return create_client(url, key)


def get_last_synced_at() -> datetime:
    sb = _supabase_client()
    if not sb:
        print("  SUPABASE creds not set, using epoch as last_synced_at", file=sys.stderr)
        return datetime.fromtimestamp(0, tz=timezone.utc)
    row = sb.table("sync_state").select("value").eq("key", "last_synced_at").execute()
    if row.data:
        return parse_dt(row.data[0]["value"])
    return datetime.fromtimestamp(0, tz=timezone.utc)


def _write_sync_log(synced_at: str, stats: dict, log_text: str = ""):
    sb = _supabase_client()
    if not sb:
        return
    has_errors = bool(stats["errors"])
    processed  = stats["new"] + stats["existing"]
    status = "failed"  if has_errors and not processed else \
             "partial" if has_errors else "success"
    trigger = os.environ.get("GITHUB_EVENT_NAME", "local")
    try:
        sb.table("sync_logs").insert({
            "synced_at":      synced_at,
            "trigger":        trigger,
            "status":         status,
            "new_count":      stats["new"],
            "existing_count": stats["existing"],
            "skipped_count":  stats["skipped"],
            "error_count":    len(stats["errors"]),
            "errors":         stats["errors"],
            "log_text":       log_text,
        }).execute()
    except Exception as e:
        print(f"  ⚠ 無法寫入 sync_logs：{e}", file=sys.stderr)


# ── Folder classification ────────────────────────────────────

def find_zhijian_file(items: list[dict]) -> dict | None:
    """在直接子項目中找含「直企」的 xlsx/xls/numbers/gsheet 檔。"""
    for item in items:
        if is_folder(item):
            continue
        if "直企" not in item["name"]:
            continue
        if item["mimeType"] == GSHEET_MIME:
            return item
        if Path(item["name"]).suffix.lower() in ZHIJIAN_EXTS:
            return item
    return None


def classify_top_folder(service, folder: dict) -> tuple[str, list[dict], dict | None, list]:
    """
    回傳 (kind, items, zhijian_item, team_entries)
    kind: 'solo' | 'group' | 'skip'
    """
    items = list_folder(service, folder["id"])
    zhijian = find_zhijian_file(items)
    if zhijian:
        return "solo", items, zhijian, []

    team_entries = []
    for item in items:
        if not is_folder(item):
            continue
        sub_items = list_folder(service, item["id"])
        sub_zhijian = find_zhijian_file(sub_items)
        if sub_zhijian:
            team_entries.append((item, sub_items, sub_zhijian))

    if team_entries:
        return "group", items, None, team_entries

    return "skip", items, None, []


# ── File sync per expedition folder ─────────────────────────

def sync_expedition_files(
    service,
    items: list[dict],
    local_dir: str,
    last_synced_at: datetime,
    is_new: bool,
    errors: list,
) -> dict:
    """
    掃描出隊資料夾內的子資料夾，下載檔案。
    回傳 {gpx_files, map_files, record_files} 各含 {drive_file_id, name, local_path}。
    下載失敗的檔案會 append 到 errors 並跳過，不中斷整個 run。
    """
    gpx_files, map_files, record_files = [], [], []

    for item in items:
        if not is_folder(item):
            continue
        folder_name = item["name"]
        sub_items = list_folder(service, item["id"])

        if folder_name in MAP_FOLDER_NAMES:
            for f in sub_items:
                if is_folder(f):
                    continue
                ext = Path(f["name"]).suffix.lower()
                if ext not in MAP_EXTS:
                    continue
                if not is_new and not (f.get("modifiedTime") and parse_dt(f["modifiedTime"]) > last_synced_at):
                    continue
                dest = MAPS_DIR / local_dir / f["name"]
                try:
                    download_file(service, f["id"], dest, f.get("modifiedTime"))
                    map_files.append({
                        "drive_file_id": f["id"],
                        "name": f["name"],
                        "local_path": str(dest),
                    })
                except Exception as e:
                    errors.append({"folder": local_dir, "file": f["name"], "stage": "download_map", "message": str(e)})

        elif folder_name in TRACK_FOLDER_NAMES:
            for f in sub_items:
                if is_folder(f):
                    continue
                ext = Path(f["name"]).suffix.lower()
                if ext not in GPX_EXTS:
                    continue
                if not is_new and not (f.get("modifiedTime") and parse_dt(f["modifiedTime"]) > last_synced_at):
                    continue
                dest = GPX_DIR / local_dir / f["name"]
                try:
                    download_file(service, f["id"], dest, f.get("modifiedTime"))
                    gpx_files.append({
                        "drive_file_id": f["id"],
                        "name": f["name"],
                        "local_path": str(dest),
                    })
                except Exception as e:
                    errors.append({"folder": local_dir, "file": f["name"], "stage": "download_gpx", "message": str(e)})

        elif folder_name in RECORD_FOLDER_NAMES:
            for f in sub_items:
                if is_folder(f):
                    continue
                mime = f["mimeType"]
                ext  = Path(f["name"]).suffix.lower()
                if not is_new and not (f.get("modifiedTime") and parse_dt(f["modifiedTime"]) > last_synced_at):
                    continue
                if mime == GDOC_MIME:
                    dest = TXT_DIR / local_dir / f"{f['name']}.txt"
                    try:
                        download_google_doc(service, f["id"], dest, f.get("modifiedTime"))
                        record_files.append({
                            "drive_file_id": f["id"],
                            "name": f"{f['name']}.txt",
                            "local_path": str(dest),
                        })
                    except Exception as e:
                        errors.append({"folder": local_dir, "file": f["name"], "stage": "download_record", "message": str(e)})
                elif mime == GSHEET_MIME:
                    pass  # TODO: Google Sheet 直企處理，之後補
                elif ext in RECORD_EXTS:
                    dest = TXT_DIR / local_dir / f["name"]
                    try:
                        download_file(service, f["id"], dest, f.get("modifiedTime"))
                        record_files.append({
                            "drive_file_id": f["id"],
                            "name": f["name"],
                            "local_path": str(dest),
                        })
                    except Exception as e:
                        errors.append({"folder": local_dir, "file": f["name"], "stage": "download_record", "message": str(e)})

    return {"gpx_files": gpx_files, "map_files": map_files, "record_files": record_files}


def build_expedition_entry(
    service,
    folder: dict,
    items: list[dict],
    zhijian: dict,
    local_dir: str,
    last_synced_at: datetime,
    is_new: bool,
    errors: list,
) -> dict:
    """下載直企檔（xlsx/xls/gsheet/numbers）並統一轉為 xlsx，回傳 expedition entry dict。
    直企下載或轉換失敗時直接 raise（caller 負責記錄到 errors）。
    """
    zhijian_mime = zhijian["mimeType"]
    zhijian_ext  = Path(zhijian["name"]).suffix.lower()
    stem = Path(zhijian["name"]).stem or zhijian["name"]

    if zhijian_mime == GSHEET_MIME:
        xlsx_dest = download_google_sheet_as_xlsx(
            service, zhijian["id"], XLSX_DIR / local_dir / f"{stem}.xlsx", zhijian.get("modifiedTime")
        )
    elif zhijian_ext == ".numbers":
        numbers_dest = XLSX_DIR / local_dir / zhijian["name"]
        download_file(service, zhijian["id"], numbers_dest, zhijian.get("modifiedTime"))
        xlsx_dest = convert_numbers_to_xlsx(numbers_dest)
    else:
        xlsx_dest = XLSX_DIR / local_dir / zhijian["name"]
        download_file(service, zhijian["id"], xlsx_dest, zhijian.get("modifiedTime"))

    files = sync_expedition_files(service, items, local_dir, last_synced_at, is_new, errors)
    return {
        "drive_folder_id": folder["id"],
        "name": folder["name"],
        "created_time": folder.get("createdTime", ""),
        "xlsx": {
            "drive_file_id": zhijian["id"],
            "name": zhijian["name"],
            "local_path": str(xlsx_dest),
        },
        **files,
    }


# ── Main ─────────────────────────────────────────────────────

class _Tee:
    def __init__(self, real):
        self._real = real
        self._buf = io.StringIO()
    def write(self, s):
        self._real.write(s)
        self._buf.write(s)
    def flush(self):
        self._real.flush()
    def getvalue(self) -> str:
        return self._buf.getvalue()


def main():
    root_id = os.environ.get("GDRIVE_ROOT_FOLDER_ID")
    if not root_id:
        print("GDRIVE_ROOT_FOLDER_ID not set", file=sys.stderr)
        sys.exit(1)

    _tee = _Tee(sys.stdout)
    sys.stdout = _tee

    service = build_service()
    last_synced_at = get_last_synced_at()
    synced_at = datetime.now(tz=timezone.utc).isoformat()

    print(f"last_synced_at: {last_synced_at.isoformat()}")

    top_folders = [f for f in list_folder(service, root_id) if is_folder(f)]

    solos: list[dict] = []
    groups: list[dict] = []
    run_stats: dict = {"new": 0, "existing": 0, "skipped": 0, "errors": []}

    for folder in top_folders:
        created = folder.get("createdTime")
        if not created:
            continue
        created_date = parse_dt(created).date()
        if created_date < IMPORT_CUTOFF_DATE:
            continue

        is_new = parse_dt(created) > last_synced_at
        print(f"\n{'[new]' if is_new else '[existing]'} {folder['name']}")

        kind, items, zhijian, team_entries = classify_top_folder(service, folder)

        if kind == "solo":
            local_dir = folder["name"]
            try:
                entry = build_expedition_entry(
                    service, folder, items, zhijian, local_dir, last_synced_at, is_new,
                    run_stats["errors"]
                )
                solos.append(entry)
                run_stats["new" if is_new else "existing"] += 1
            except Exception as e:
                run_stats["errors"].append({
                    "folder": folder["name"],
                    "stage": "build",
                    "message": str(e),
                })
                print(f"  ✗ {e}")

        elif kind == "group":
            group_entry: dict = {
                "drive_folder_id": folder["id"],
                "name": folder["name"],
                "created_time": folder.get("createdTime", ""),
                "teams": [],
            }
            for team_folder, team_items, team_zhijian in team_entries:
                team_created = team_folder.get("createdTime", "")
                team_is_new = bool(team_created) and parse_dt(team_created) > last_synced_at
                local_dir = f"{folder['name']}/{team_folder['name']}"
                print(f"  team: {team_folder['name']}")
                try:
                    team_entry = build_expedition_entry(
                        service, team_folder, team_items, team_zhijian,
                        local_dir, last_synced_at, team_is_new,
                        run_stats["errors"]
                    )
                    group_entry["teams"].append(team_entry)
                    run_stats["new" if team_is_new else "existing"] += 1
                except Exception as e:
                    run_stats["errors"].append({
                        "folder": local_dir,
                        "stage": "build",
                        "message": str(e),
                    })
                    print(f"  ✗ {e}")

            if group_entry["teams"]:
                groups.append(group_entry)

        else:
            print(f"  skipped (no 直企 xlsx found)")
            run_stats["skipped"] += 1

    RAW_DIR.mkdir(parents=True, exist_ok=True)
    META_PATH.write_text(json.dumps({
        "synced_at": synced_at,
        "solos": solos,
        "groups": groups,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    sys.stdout = _tee._real
    _write_sync_log(synced_at, run_stats, log_text=_tee.getvalue())
    print(f"\nsync complete → {META_PATH}")


if __name__ == "__main__":
    main()
