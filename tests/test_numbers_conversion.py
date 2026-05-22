"""
測試 convert_numbers_to_xlsx 輸出是否能被 capture_sheet 截圖。

前置作業（只需一次）：
  1. macOS 開 Numbers，sheet 名稱改為「直企P1(列印)」
  2. 新增第二個 sheet「直企P2(列印)」
  3. P1 的 D2 填出隊名，C3/C4 填日期，F3/F4 填地名，C17 填隊長名
  4. 存成 .numbers 格式，放到 tests/fixtures/sample.numbers

若 sample.numbers 不存在，所有測試自動跳過。
"""
import shutil
import sys
import pytest
from pathlib import Path
from PIL import Image

SAMPLE = Path(__file__).parent / "fixtures/sample.numbers"


@pytest.fixture(scope="module")
def sd():
    sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))
    import sync_drive
    return sync_drive


@pytest.mark.skipif(not SAMPLE.exists(), reason="需要 tests/fixtures/sample.numbers")
@pytest.mark.timeout(60)
def test_numbers_produces_xlsx(sd, tmp_path):
    src = tmp_path / "sample.numbers"
    shutil.copy(SAMPLE, src)
    result = sd.convert_numbers_to_xlsx(src)
    assert result is not None, "convert_numbers_to_xlsx 回傳 None"
    assert result.exists(), f"轉換後 xlsx 不存在：{result}"
    assert result.suffix.lower() == ".xlsx"


@pytest.mark.skipif(not SAMPLE.exists(), reason="需要 tests/fixtures/sample.numbers")
@pytest.mark.timeout(60)
def test_numbers_xlsx_has_p1_sheet(sd, tmp_path):
    src = tmp_path / "sample.numbers"
    shutil.copy(SAMPLE, src)
    xlsx = sd.convert_numbers_to_xlsx(src)
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    p1_candidates = ["直企P1(列印)", "直企列印 P1"]
    found = any(name in wb.sheetnames for name in p1_candidates)
    assert found, f"找不到 P1 sheet，實際 sheets：{wb.sheetnames}"


@pytest.mark.skipif(not SAMPLE.exists(), reason="需要 tests/fixtures/sample.numbers")
@pytest.mark.timeout(120)
def test_numbers_capture_p1(sd, nm, tmp_path):
    """端對端：numbers → xlsx → capture_sheet → PNG"""
    import openpyxl
    src = tmp_path / "sample.numbers"
    shutil.copy(SAMPLE, src)
    xlsx = sd.convert_numbers_to_xlsx(src)
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    p1 = nm.find_sheet(wb, nm.P1_NAMES)
    assert p1 is not None, f"轉換後找不到 P1 sheet，實際 sheets：{wb.sheetnames}"
    out = tmp_path / "p1.png"
    nm.capture_sheet(xlsx, p1, None, out)
    assert out.exists(), "截圖未生成"
    img = Image.open(out)
    assert img.format == "PNG"
    assert img.width >= 100 and img.height >= 100, f"截圖尺寸異常：{img.width}x{img.height}"
