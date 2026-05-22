"""
測試 Google Sheet 匯出的 xlsx 能被 capture_sheet 截圖。

fixture 是從 Google Sheets 網頁手動「下載為 xlsx」的檔案，
放在 tests/fixtures/茶茶牙頓出西都驕溪_all in one直企.xlsx。

若要測試完整的 Drive API 下載流程（download_google_sheet_as_xlsx），
需設定環境變數 GDRIVE_CREDENTIALS_JSON 和 GDRIVE_TEST_SHEET_ID。
"""
import json
import os
import sys
import pytest
from pathlib import Path
from PIL import Image

GSHEET_XLSX = Path(__file__).parent / "fixtures/茶茶牙頓出西都驕溪_all in one直企.xlsx"

HAS_CREDS    = bool(os.environ.get("GDRIVE_CREDENTIALS_JSON"))
HAS_SHEET_ID = bool(os.environ.get("GDRIVE_TEST_SHEET_ID"))
INTEGRATION_SKIP = "需要 GDRIVE_CREDENTIALS_JSON 和 GDRIVE_TEST_SHEET_ID 環境變數"


# ── Local fixture tests（不需要 env var）─────────────────────

@pytest.mark.skipif(not GSHEET_XLSX.exists(), reason="需要 tests/fixtures/ 的 gsheet xlsx")
def test_gsheet_xlsx_has_p1_sheet(nm):
    import openpyxl
    wb = openpyxl.load_workbook(GSHEET_XLSX, data_only=True)
    p1 = nm.find_sheet(wb, nm.P1_NAMES)
    assert p1 is not None, f"找不到 P1 sheet，實際 sheets：{wb.sheetnames}"


@pytest.mark.skipif(not GSHEET_XLSX.exists(), reason="需要 tests/fixtures/ 的 gsheet xlsx")
@pytest.mark.timeout(60)
def test_gsheet_capture_p1(nm, tmp_path):
    import openpyxl
    wb = openpyxl.load_workbook(GSHEET_XLSX, data_only=True)
    p1 = nm.find_sheet(wb, nm.P1_NAMES)
    assert p1 is not None, f"找不到 P1 sheet，實際 sheets：{wb.sheetnames}"
    out = tmp_path / "p1.png"
    nm.capture_sheet(GSHEET_XLSX, p1, None, out)
    assert out.exists(), "截圖未生成"
    img = Image.open(out)
    assert img.format == "PNG"
    assert img.width >= 100 and img.height >= 100, f"截圖尺寸異常：{img.width}x{img.height}"


@pytest.mark.skipif(not GSHEET_XLSX.exists(), reason="需要 tests/fixtures/ 的 gsheet xlsx")
@pytest.mark.timeout(60)
def test_gsheet_capture_p2(nm, tmp_path):
    import openpyxl
    wb = openpyxl.load_workbook(GSHEET_XLSX, data_only=True)
    p2 = nm.find_sheet(wb, nm.P2_NAMES)
    if not p2:
        pytest.skip("此 xlsx 無 P2 sheet")
    out = tmp_path / "p2.png"
    nm.capture_sheet(GSHEET_XLSX, p2, "B2:O12", out)
    assert out.exists(), "截圖未生成"
    img = Image.open(out)
    assert img.format == "PNG"
    assert img.width >= 100 and img.height >= 100


# ── Integration tests（需要 Drive API env var）───────────────

@pytest.fixture(scope="module")
def real_service():
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    creds = Credentials.from_service_account_info(
        json.loads(os.environ["GDRIVE_CREDENTIALS_JSON"]),
        scopes=["https://www.googleapis.com/auth/drive.readonly"],
    )
    return build("drive", "v3", credentials=creds)


@pytest.fixture(scope="module")
def sd():
    sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))
    import sync_drive
    return sync_drive


@pytest.mark.skipif(not (HAS_CREDS and HAS_SHEET_ID), reason=INTEGRATION_SKIP)
@pytest.mark.timeout(120)
def test_gsheet_api_capture_p1(sd, nm, real_service, tmp_path):
    """端對端：Drive API 下載 → xlsx → capture_sheet → PNG"""
    import openpyxl
    dest = tmp_path / "exported.xlsx"
    result = sd.download_google_sheet_as_xlsx(
        real_service, os.environ["GDRIVE_TEST_SHEET_ID"], dest
    )
    assert result.exists(), "Drive API 匯出失敗"
    wb = openpyxl.load_workbook(result, data_only=True)
    p1 = nm.find_sheet(wb, nm.P1_NAMES)
    assert p1 is not None, f"匯出的 xlsx 找不到 P1 sheet，實際 sheets：{wb.sheetnames}"
    out = tmp_path / "p1.png"
    nm.capture_sheet(result, p1, None, out)
    assert out.exists(), "截圖未生成"
    img = Image.open(out)
    assert img.width >= 100 and img.height >= 100
